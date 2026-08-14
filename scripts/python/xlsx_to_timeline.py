#!/usr/bin/env python3
"""
Turn a corrected review workbook into the app's timeline file.

Whatever is typed in "fixed_ms" wins over the proposal; empty cells keep it.
The result is validated before it is written — every word must be on the page
and the marks must run forwards — because a timeline that is quietly wrong is
worse than one that is missing.

    python scripts/python/xlsx_to_timeline.py --xlsx review_553.xlsx
    node scripts/import-timeline.mjs public/timeline/553.json
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[2]
ANNOTATIONS = REPO / "public" / "annotations"
TIMELINES = REPO / "public" / "timeline"

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--xlsx", type=Path, required=True)
    parser.add_argument("--out", type=Path, default=None)
    args = parser.parse_args()

    sheet = load_workbook(args.xlsx, data_only=True).active
    head = [c.value for c in sheet[1]]
    meta = dict(zip(head[::2], head[1::2]))
    page = int(meta["page"])
    start_ms = int(meta.get("start_ms") or 0)
    duration_ms = int(meta.get("duration_ms") or 0)

    raw = json.loads((ANNOTATIONS / f"{page}.json").read_text(encoding="utf8"))
    known = {int(w["id"]) for w in (raw if isinstance(raw, list) else raw.get("words", [])) if w.get("id") is not None}

    events, problems = [], []
    for row in sheet.iter_rows(min_row=4, values_only=True):
        if not row or row[0] is None:
            continue
        word_id = int(row[0])
        proposed, fixed = row[4], row[7]
        ms = int(fixed if fixed not in (None, "") else proposed)
        if word_id not in known:
            problems.append(f"word {word_id} is not on page {page}")
        events.append({"t": max(0, ms - start_ms), "w": word_id})

    events.sort(key=lambda e: e["t"])
    for a, b in zip(events, events[1:]):
        if b["t"] == a["t"]:
            problems.append(f"words {a['w']} and {b['w']} share a millisecond")

    if problems:
        for p in problems[:10]:
            print(f"✗ {p}")
        print("nothing written")
        return 1

    out = args.out or (TIMELINES / f"{page}.json")
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(
        json.dumps(
            {
                "page": page,
                "audio": meta.get("audio"),
                "start": start_ms,
                "duration": duration_ms,
                "events": events,
            }
        ),
        encoding="utf8",
    )
    print(f"wrote {out} — {len(events)} words, {start_ms} ms → {start_ms + duration_ms} ms")
    return 0


if __name__ == "__main__":
    sys.exit(main())
