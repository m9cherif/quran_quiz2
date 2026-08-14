#!/usr/bin/env python3
"""
Propose a page's word times from the recording, for a human to correct.

Semi-automatic, and honest about which half is which. Finding where the reciter
*stops* needs no model at all: a pause is a long dip in energy. Telling a word
onset from a syllable onset inside a phrase does need one, and this has none —
page 553 offers about 300 energy onsets for 118 words, and nothing in the
signal says which 118 are the words.

So it does the part it can do well and marks the rest as a draft. Measured
against the hand-made timeline for page 553:

    ayah boundaries   73-412 ms   (pinned to the reciter's long pauses)
    every word        median 1.1 s, p90 2.4 s

Which is to say: the ayah rows are right, the word rows are a starting point
that still has to be corrected by ear. What it buys you over tapping from
scratch is that the page is already in order, already split at every pause, and
an error can never travel past the end of its ayah. For word-level precision
without a human, use align_timeline.py instead.

The words come from the annotation workbook (the xlsx files in the data repo's
annotation/ folder) or from public/annotations/{page}.json.

    pip install -r scripts/python/requirements-semiauto.txt

    python scripts/python/propose_timeline.py \\
        --page 553 --audio ~/audio/062.mp3 --start-ms 6500 --out review_553.xlsx

Open review_553.xlsx, listen, and type a corrected millisecond in the "fixed"
column for any word that is off. Leave the rest empty. Then:

    python scripts/python/xlsx_to_timeline.py --xlsx review_553.xlsx
    node scripts/import-timeline.mjs public/timeline/553.json
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path

import numpy as np
import soundfile as sf
from openpyxl import Workbook, load_workbook

REPO = Path(__file__).resolve().parents[2]
ANNOTATIONS = REPO / "public" / "annotations"

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

HOP_MS = 10
TASHKEEL = re.compile("[ؐ-ًؚ-ٰٟۖ-ۭـ]")


def strip_tashkeel(text: str) -> str:
    return TASHKEEL.sub("", unicodedata.normalize("NFC", str(text))).strip()


# ---------------------------------------------------------------- the words


def words_from_json(page: int) -> list[dict]:
    raw = json.loads((ANNOTATIONS / f"{page}.json").read_text(encoding="utf8"))
    rows = raw if isinstance(raw, list) else raw.get("words", [])
    return [
        {"id": int(w["id"]), "text": str(w["text"]).strip(), "aya": w.get("aya")}
        for w in rows
        if w.get("id") is not None and str(w.get("text", "")).strip()
    ]


def words_from_xlsx(path: Path) -> list[dict]:
    """The annotation workbook, read the same way the JS importer reads it."""
    book = load_workbook(path, read_only=True, data_only=True)
    sheet = book["Mots"] if "Mots" in book.sheetnames else book[book.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    index = {name: i for i, name in enumerate(header)}

    def cell(row, *names):
        for name in names:
            if name in index and index[name] < len(row):
                value = row[index[name]]
                if value not in (None, ""):
                    return value
        return None

    out = []
    for row in rows:
        text = strip_tashkeel(cell(row, "kalima_text_emlaey") or "") or strip_tashkeel(
            cell(row, "kalima_text", "text") or ""
        )
        word_id = cell(row, "id")
        if not text or word_id is None:
            continue
        out.append({"id": int(word_id), "text": text, "aya": cell(row, "aya_no", "aya")})
    return out


# ---------------------------------------------------------------- the audio


def energy(audio: Path, start_ms: int, end_ms: int | None) -> tuple[np.ndarray, int]:
    """RMS every 10 ms, plus the offset it starts at."""
    data, sr = sf.read(str(audio), dtype="float32", always_2d=True)
    mono = data.mean(axis=1)
    first = int(start_ms / 1000 * sr)
    last = int(end_ms / 1000 * sr) if end_ms else len(mono)
    mono = mono[first:last]
    hop = max(1, int(sr * HOP_MS / 1000))
    frames = len(mono) // hop
    if frames == 0:
        raise SystemExit("that slice of audio is empty")
    rms = np.sqrt((mono[: frames * hop].reshape(frames, hop) ** 2).mean(axis=1))
    return rms, start_ms


def find_phrases(
    rms: np.ndarray, min_silence_ms: int, floor_percentile: float, min_phrase_ms: int = 220
) -> list[tuple[int, int]]:
    """
    Stretches of speech, split at every real pause.

    The threshold is taken from the recording itself rather than fixed: a quiet
    recording and a loud one have nothing in common in absolute terms, but in
    both the pauses sit near the bottom of the distribution.
    """
    floor = np.percentile(rms, floor_percentile)
    loud = np.percentile(rms, 85)
    threshold = floor + (loud - floor) * 0.12
    quiet = rms < threshold

    phrases: list[tuple[int, int]] = []
    run_start = None
    min_frames = max(1, min_silence_ms // HOP_MS)
    silence_run = 0
    for i, is_quiet in enumerate(quiet):
        if is_quiet:
            silence_run += 1
            if silence_run == min_frames and run_start is not None:
                phrases.append((run_start, i - min_frames + 1))
                run_start = None
        else:
            silence_run = 0
            if run_start is None:
                run_start = i
    if run_start is not None:
        phrases.append((run_start, len(quiet)))

    # A burst too short to be a word is breath, a reverb tail, or one word cut
    # in two by its own silence — a stop consonant does that. Left alone these
    # take a word each and the whole page runs early, so they are folded into
    # the phrase before them.
    min_frames_phrase = max(1, min_phrase_ms // HOP_MS)
    merged: list[tuple[int, int]] = []
    for start, end in phrases:
        if end - start >= min_frames_phrase:
            merged.append((start, end))
        elif merged:
            merged[-1] = (merged[-1][0], end)
    return merged


def tune_phrases(rms: np.ndarray, word_count: int, floor_percentile: float) -> tuple[list[tuple[int, int]], int]:
    """
    Find the pause length that suits this reciter.

    How much a reciter pauses between words is a property of the recitation, not
    a constant: a slow teaching recitation stops after almost every word, while
    a flowing one stops only at the waqf. So rather than fix a threshold, try a
    range and keep the one that carves the audio into as many phrases as
    possible without exceeding the number of words — one phrase per word is the
    ideal case, and it makes the proposal exact.
    """
    best: tuple[list[tuple[int, int]], int] = ([], 0)
    for min_silence in range(60, 420, 20):
        phrases = find_phrases(rms, min_silence, floor_percentile)
        if len(phrases) <= word_count and len(phrases) > len(best[0]):
            best = (phrases, min_silence)
    if not best[0]:  # every setting over-segments: take the coarsest
        best = (find_phrases(rms, 400, floor_percentile), 400)
    return best


def ayah_anchors(rms: np.ndarray, words: list[dict], long_silence_ms: int = 700) -> dict[int, int]:
    """
    Pin the first word of each ayah to where the reciter resumes.

    This is the one thing energy alone gets right. A reciter stops at the end of
    an ayah far longer than between words, so the ends of the long silences are
    the ayah boundaries — measured against a hand-made timeline for page 553,
    every ayah start landed within about 300 ms. Pinning them means an error
    inside one ayah cannot leak into the next, which is what turns a useless
    draft into one worth correcting.
    """
    floor = np.percentile(rms, 20)
    loud = np.percentile(rms, 85)
    quiet = rms < floor + (loud - floor) * 0.12
    resumes, run = [], None
    for i, is_quiet in enumerate(quiet):
        if is_quiet and run is None:
            run = i
        elif not is_quiet and run is not None:
            if i - run >= long_silence_ms // HOP_MS:
                resumes.append(i)
            run = None
    if not resumes:
        return {}

    starts = [i for i, w in enumerate(words) if i == 0 or w.get("aya") != words[i - 1].get("aya")]
    if len(starts) < 2:
        return {}

    # Match ayah starts to resumptions in order, keeping the ones that fall
    # near where an even spread would put them.
    anchors = {0: 0}
    available = list(resumes)
    for index in starts[1:]:
        expected = len(rms) * index / len(words)
        near = [r for r in available if abs(r - expected) < len(rms) * 0.12]
        if not near:
            continue
        chosen = min(near, key=lambda r: abs(r - expected))
        if all(chosen > a for a in anchors.values()):
            anchors[index] = chosen
            available = [r for r in available if r > chosen]
    return anchors


def local_minima(rms: np.ndarray) -> np.ndarray:
    """Frames quieter than both neighbours — candidate word boundaries."""
    smooth = np.convolve(rms, np.ones(3) / 3, mode="same")
    return np.where((smooth[1:-1] <= smooth[:-2]) & (smooth[1:-1] <= smooth[2:]))[0] + 1


def partition(
    phrases: list[tuple[int, int]], weights: np.ndarray, rate: float, max_per_phrase: int = 8
) -> list[int]:
    """
    How many words each phrase holds.

    Deciding phrase by phrase is what ruins this: choose one phrase's share
    greedily and every later phrase inherits the mistake, so the page ends up a
    word or two out and stays that way. The split has to be chosen as a whole —
    which is a shortest-path problem, and small enough to solve exactly.

    The cost of giving a phrase a run of words is how badly their expected
    duration misses the phrase's real one, at this reciter's rate.
    """
    words = len(weights)
    cumulative = np.concatenate([[0.0], np.cumsum(weights)])
    best = np.full((len(phrases) + 1, words + 1), np.inf)
    take = np.zeros((len(phrases) + 1, words + 1), dtype=int)
    best[0, 0] = 0.0

    for p, (start, end) in enumerate(phrases):
        duration = end - start
        left_after = len(phrases) - p - 1
        for i in range(words + 1):
            if not np.isfinite(best[p, i]):
                continue
            # Leave at least one word for every phrase still to come.
            for k in range(1, min(max_per_phrase, words - i - left_after) + 1):
                expected = (cumulative[i + k] - cumulative[i]) * rate
                cost = best[p, i] + abs(expected - duration)
                if cost < best[p + 1, i + k]:
                    best[p + 1, i + k] = cost
                    take[p + 1, i + k] = k

    counts: list[int] = []
    i = words
    for p in range(len(phrases), 0, -1):
        k = take[p, i]
        counts.append(k)
        i -= k
    return counts[::-1]


# ------------------------------------------------------------- putting them together


def propose(
    rms: np.ndarray,
    words: list[dict],
    phrases: list[tuple[int, int]],
    offset_ms: int,
    anchors: dict[int, int] | None = None,
) -> list[dict]:
    """
    Place every word in the phrases, in order.

    Words are spread over the phrases by weight — a long word takes longer to
    say than a short one — and each boundary is then pulled to the quietest
    frame near where it was expected. Pauses anchor the whole thing, so an error
    inside one phrase cannot leak into the next.
    """
    weights = np.array([max(1, len(w["text"])) for w in words], dtype=float)
    speech = sum(b - a for a, b in phrases)
    if speech == 0 or not phrases:
        raise SystemExit("no speech found — check --start-ms/--end-ms and --floor")

    # Ayah by ayah, when the words say which ayah they belong to. Each ayah is
    # placed inside its own stretch of audio at its own rate, so a word misplaced
    # in one cannot push the rest of the page along with it.
    if anchors and len(anchors) > 1:
        out: list[dict] = []
        bounds = sorted(anchors.items())
        for (word_index, frame), nxt in zip(bounds, bounds[1:] + [(len(words), len(rms))]):
            span = [(max(a, frame), min(b, nxt[1])) for a, b in phrases if b > frame and a < nxt[1]]
            span = [(a, b) for a, b in span if b > a] or [(frame, nxt[1])]
            piece = propose(rms, words[word_index : nxt[0]], span, offset_ms, anchors=None)
            out.extend(piece)
        return out

    # How many words each phrase gets. Not proportional to the phrase's share of
    # the recording — that ignores which words they are — but spent at the
    # reciter's own rate: a phrase lasting a second and a half buys about a
    # second and a half of syllables.
    # Every phrase must hold at least one word, so more phrases than words has
    # no solution: fold the shortest into its neighbour until it does. Those are
    # the least word-like anyway.
    phrases = list(phrases)
    while len(phrases) > len(words):
        i = min(range(len(phrases)), key=lambda j: phrases[j][1] - phrases[j][0])
        j = i - 1 if i else 1
        a, b = min(phrases[i][0], phrases[j][0]), max(phrases[i][1], phrases[j][1])
        phrases[min(i, j)] = (a, b)
        phrases.pop(max(i, j))

    rate = speech / weights.sum()  # frames per letter
    counts = partition(phrases, weights, rate)

    minima = local_minima(rms)
    out = []
    cursor = 0
    for (start, end), count in zip(phrases, counts):
        chunk = words[cursor : cursor + count]
        if not chunk:
            continue
        w = weights[cursor : cursor + count]
        # Expected onsets inside the phrase, by cumulative weight.
        fractions = np.concatenate([[0.0], np.cumsum(w)[:-1] / w.sum()])
        expected = start + fractions * (end - start)

        previous = start
        for i, (word, frame) in enumerate(zip(chunk, expected)):
            if i == 0:
                placed = start  # a phrase starts where the sound starts
            else:
                span = (end - start) * (w[i] / w.sum())
                window = max(3, int(span * 0.45))
                near = minima[(minima >= frame - window) & (minima <= frame + window)]
                placed = int(near[np.argmin(rms[near])]) if len(near) else int(round(frame))
                placed = max(previous + 2, min(placed, end - 2))
            previous = placed
            out.append(
                {
                    "id": word["id"],
                    "text": word["text"],
                    "aya": word.get("aya"),
                    "ms": offset_ms + placed * HOP_MS,
                    "phrase": len(out) and out[-1]["phrase"] or 0,
                    "anchored": i == 0,
                }
            )
        cursor += count
    return out


def write_review(rows: list[dict], path: Path, page: int, audio: str, start_ms: int, duration_ms: int) -> None:
    book = Workbook()
    sheet = book.active
    sheet.title = "timeline"
    sheet.append(["page", page, "audio", audio, "start_ms", start_ms, "duration_ms", duration_ms])
    sheet.append([])
    sheet.append(["id", "word", "aya", "phrase", "proposed_ms", "gap_ms", "anchored", "fixed_ms"])
    for i, row in enumerate(rows):
        gap = row["ms"] - rows[i - 1]["ms"] if i else 0
        sheet.append(
            [
                row["id"],
                row["text"],
                row["aya"],
                row["phrase"],
                row["ms"],
                gap,
                "yes" if row["anchored"] else "",
                None,
            ]
        )
    sheet.column_dimensions["B"].width = 18
    sheet.freeze_panes = "A4"
    book.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--page", type=int, required=True)
    parser.add_argument("--audio", type=Path, required=True)
    parser.add_argument("--xlsx-words", type=Path, help="annotation workbook; defaults to public/annotations/{page}.json")
    parser.add_argument("--start-ms", type=int, default=0, help="where the page begins in the recording")
    parser.add_argument("--end-ms", type=int, default=None)
    parser.add_argument("--min-silence-ms", type=int, default=0,
                        help="what counts as a pause; 0 lets the tool find it")
    parser.add_argument("--floor", type=float, default=20.0, help="percentile taken as the noise floor")
    parser.add_argument("--compare", type=Path,
                        help="an existing timeline to measure the proposal against")
    parser.add_argument("--no-anchors", action="store_true",
                        help="do not pin ayah starts to the long pauses")
    parser.add_argument("--out", type=Path, default=None)
    args = parser.parse_args()

    words = words_from_xlsx(args.xlsx_words) if args.xlsx_words else words_from_json(args.page)
    print(f"page {args.page}: {len(words)} words")

    rms, offset = energy(args.audio, args.start_ms, args.end_ms)
    if args.min_silence_ms:
        phrases, silence = find_phrases(rms, args.min_silence_ms, args.floor), args.min_silence_ms
    else:
        phrases, silence = tune_phrases(rms, len(words), args.floor)
    print(
        f"audio: {len(rms) * HOP_MS / 1000:.1f}s from {args.start_ms} ms · "
        f"{len(phrases)} phrases at a {silence} ms pause"
    )
    if len(phrases) > len(words):
        print("  ! more phrases than words — raise --min-silence-ms")
    elif len(phrases) >= len(words) * 0.9:
        print("  · nearly one phrase per word — this reciter stops between words, "
              "which is the best case for this method")

    anchors = {} if args.no_anchors else ayah_anchors(rms, words)
    if anchors:
        print(f"  · {len(anchors)} ayah boundaries pinned to the reciter's long pauses")
    rows = propose(rms, words, phrases, offset, anchors=anchors)
    if args.compare:
        known = json.loads(args.compare.read_text(encoding="utf8"))
        truth = {e["w"]: known.get("start", 0) + e["t"] for e in known["events"]}
        errors = sorted(abs(r["ms"] - truth[r["id"]]) for r in rows if r["id"] in truth)
        if errors:
            print(f"against {args.compare.name}: median {errors[len(errors) // 2]} ms, "
                  f"p90 {errors[int(len(errors) * 0.9)]} ms, worst {errors[-1]} ms "
                  f"({sum(e <= 500 for e in errors)}/{len(errors)} within half a second)")

    out = args.out or Path(f"review_{args.page}.xlsx")
    write_review(rows, out, args.page, args.audio.name, args.start_ms, len(rms) * HOP_MS)
    print(f"wrote {out} — {len(rows)} rows, {sum(r['anchored'] for r in rows)} of them anchored on a pause")
    print(f"correct what is off in the 'fixed_ms' column, then:\n"
          f"  python scripts/python/xlsx_to_timeline.py --xlsx {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
